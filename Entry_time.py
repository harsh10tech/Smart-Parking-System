import cv2 as cv
import pytesseract
import os
import time
import class_thres
from PIL import Image
from datetime import date
import openpyxl

#calling of needed software
pytesseract.pytesseract.tesseract_cmd=(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe")
yes = '0'

sence = input('Type 0 to enter: ')

if sence == yes :
    
   
    date=date.today().strftime('%d-%B-%Y')
    x= time.localtime()
    entry_time = time.strftime("%H:%M:%S", x)
    
    #will create new text file for every new day with their title as date
    new_day = date+'.txt'
    
    #Capturing the image of the plate
    num_plate = cv.VideoCapture(0) 
    check,plate = num_plate.read() #reading the image
    
    gray = cv.cvtColor(plate, cv.COLOR_BGR2GRAY) #changing image into grayscale
    
    cv.imshow('number_plate',plate) #opening a window to show the image
    
    #A class for thresholld and creating temporary file
    gray = class_thres.thres.thres2(gray) 
    temp = "{}.png".format(os.getpid())
    cv.imwrite(temp,gray)
    
    #Extracting the text from the image in string form
    number = pytesseract.image_to_string(Image.open(temp))
    
    #deleting the temporary file
    os.remove(temp)
    
    time.sleep(3)
    
    data = '\n' + number + '   ' + entry_time
    
    #A text file to keep permanent record of the entries for future record
    entry = open(new_day,"a")

    entry.write(data)
    entry.close()
    

    wb= openpyxl.load_workbook(r'C:/Users/harsh/Documents/MP- Smart Parking/Temp record.xlsx')
    
    sheet1 = wb.active
    
    max_col = sheet1.max_row
    x=0
    #entry of Number paper and entry time in exel sheet
    for i in range(1, max_col + 1): 
        s= sheet1.cell(row=i,column=1)
        t= sheet1.cell(row=i,column=2)
        if s.value==x:
            s.value = number
            t.value = entry_time
            break
    
    
    wb.save(r'C:/Users/harsh/Documents/MP- Smart Parking/Temp record.xlsx')
    
    
cv.waitKey(0)

num_plate.release()

cv.destroyAllWindows()


    

