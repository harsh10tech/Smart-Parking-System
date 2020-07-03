import cv2 as cv
import pytesseract
import os
import time
import class_thres
from PIL import Image
from datetime import date
import openpyxl

pytesseract.pytesseract.tesseract_cmd=(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe")
yes = '0'

sence = input('Type 0 to enter: ')

if sence == yes :
    
    date=date.today().strftime('%d-%B-%Y')
    x= time.localtime()
    exit_time = time.strftime("%H:%M:%S", x)
    
    #will create the existing text file with same date
    new_day = date+'.txt'
    
    num_plate = cv.VideoCapture(0)
    check,plate = num_plate.read()
    
    gray = cv.cvtColor(plate, cv.COLOR_BGR2GRAY)
    
    cv.imshow('number_plate',plate)
    
    gray = class_thres.thres.thres2(gray)
    temp = "{}.png".format(os.getpid())
    cv.imwrite(temp,gray)
    
    number = pytesseract.image_to_string(Image.open(temp))
    
    os.remove(temp)
    
    time.sleep(3)
    
    #data = number + '   ' + entry_time
    
    #A text file to keep permanent record of the entries for future record
    exit_t = open(new_day,"r+")
    lines = exit_t.readlines()
    for n in lines:
        if number in lines:
            offset=exit_t.tell()
            exit_t.seek(offset)
            exit_t.write(exit_time)
            break
        else:
            cv.waitKey(0)
            num_plate.release()
            cv.destroyAllWindows()
            
    exit_t.close()
    

    wb= openpyxl.load_workbook(r'C:/Users/harsh/Documents/MP SEM-2/Temp record.xlsx')
    
    sheet1 = wb.active
    
    max_col = sheet1.max_row
    
    #calculation of time during exit
    for i in range(1, max_col + 1): 
        s= sheet1.cell(row=i,column=1)
        t= sheet1.cell(row=i,column=3)
        d= sheet1.cell(row=i,column=4)
        if s.value==number:
            s.value = '0'
            t.value = exit_time
            d.value = '=TEXT(C'+ i+'-B'+i+', "mm")'
            duration= d.value
            break
    
    wb.save(r'C:/Users/harsh/Documents/MP- Smart Parking/Temp record.xlsx')
    
print('Total duration is ',duration)

print('/nTotal cost is ₹', duration*2)    
    
cv.waitKey(0)

num_plate.release()

cv.destroyAllWindows()